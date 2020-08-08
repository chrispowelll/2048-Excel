# 2048 Excel

import design
import controls
import os
import sys
import random
import openpyxl
import xlwings as xw
import time

# If Windows is used, import WConio2
if os.name == 'nt':
    import WConio2
    userOS = "Windows"


def main():
    while True:
        # Original file input
        wbFilename = input("Enter your Excel file path: ")

        # Check if user wants to quit
        if wbFilename in ('quit', 'Quit', 'QUIT'):
            sys.exit()

        # Check if file exists
        elif os.path.exists(wbFilename):
            # Time limit
            while True:
                timeLimitOption = input("Would you like to set a time limit? (Y/N) ").lower()
                if timeLimitOption in ['yes', 'y']:
                    timeLimitExists = True
                    break
                elif timeLimitOption in ['no', 'n']:
                    timeLimitExists = False
                    break
                else:
                    print("Please try again. Enter Yes or No")
            if timeLimitExists:
                while True:
                    timeLimit = input("How many minutes would you like to play for? ")
                    try:
                        if int(timeLimit) > 0:
                            break
                        else:
                            print("Please enter a number greater than 0")
                    except:
                        print("Please enter an integer")

            # Create 2048 workbook
            design.create2048Workbook(wbFilename)

            # Check if they want to load saved game
            while True:
                loadSaveOption = input("Would you like to load a saved game? (Y/N) ").lower()
                if loadSaveOption in ['yes', 'y']:
                    break
                elif loadSaveOption in ['no', 'n']:
                    # Create workbook
                    design.createValuesWorkbook()
                    controls.newGame()
                    break
                else:
                    print("Please try again. Enter Yes or No")

            # Open original workbook and 2048 copy
            updateValues()

            # Start game clock
            if timeLimitExists:
                startTime = time.time()
                # Calculate when timer ends
                endTime = (int(timeLimit)*60) + startTime

            # The Game
            while True:
                # Check if time is up
                if timeLimitExists:
                    currentTime = time.time()
                    if currentTime >= endTime:
                        print("You have now played for " + timeLimit + " minutes")
                        # Option to continue playing
                        while True:
                            continuePlaying = input("Would you like to continue playing? (Y/N) ").lower()
                            if continuePlaying in ['yes', 'y']:
                                timeLimitExists = False
                                break
                            elif continuePlaying in ['no', 'n']:
                                # Quit game
                                os.startfile(wbFilename)
                                sys.exit()
                            else:
                                print("Please try again. Enter Yes or No")

                if userOS == "Windows":
                    move = WConio2.getkey()
                else:
                    move = input()

                if move == 'w':
                    controls.move("up")
                if move == 's':
                    controls.move("down")
                if move == 'a':
                    controls.move("left")
                if move == 'd':
                    controls.move("right")
                if move == 'r':
                    controls.newGame()
                    updateValues()
                if move == 'q':
                    # Open original workbook
                    os.startfile(wbFilename)

                    # Close program
                    sys.exit()

        # If file doesn't exist
        else:
            print('Invalid file location. Please try again. You can also enter QUIT to exit')


def getNewValue():
    # 90% chance 2 is generated, 10% chance 4 is generated
    numberPicker = random.randint(0, 10)
    if numberPicker >= 9:
        return 4
    else:
        return 2


def updateValues():
    gameCells = ['A2', 'A3', 'A4', 'A5', 'B2', 'B3', 'B4', 'B5', 'C2', 'C3', 'C4', 'C5', 'D2', 'D3', 'D4', 'D5']

    # Open 2048 workbook
    wb2048 = xw.Book("2048.xlsx")
    ws2048 = wb2048.sheets["2048"]

    # Open values workbook
    wbValues = openpyxl.load_workbook("values.xlsx", read_only=False)
    wsValues = wbValues.active

    # Update moves made
    ws2048.range('D1').value = wsValues['D1'].value

    # Update game status
    ws2048.range('A6').value = wsValues['A6'].value

    # Update values
    for i in gameCells:
        # Only show values greater than 0
        if wsValues[i].value > 0:
            ws2048.range(i).value = wsValues[i].value
        else:
            ws2048.range(i).value = ""


if __name__ == "__main__":
    main()
