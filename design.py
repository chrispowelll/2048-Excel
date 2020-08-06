# Creates 2048 worksheet and design

import openpyxl
from openpyxl.styles import Font, Border, Side


def create2048Workbook(wbFilename):
    # Open Excel workbook
    wb = openpyxl.load_workbook(wbFilename)

    # Create 2048 worksheet
    ws2048 = wb.copy_worksheet(wb.active)
    ws2048.title = "2048"

    # Delete other sheets
    sheets = wb.get_sheet_names()
    sheetName = wb.sheetnames
    for i in range(0, len(sheets)):
        if sheets[i] != "2048":
            wb.remove(wb[sheetName[i]])

    # Headings fonts
    for i in ['A1', 'B1', 'C1', 'D1', 'E1']:
        ws2048[i].font = Font(name='Verdana', size=10, bold=True)
    # Games fonts and border
    for i in ['A2', 'A3', 'A4', 'A5', 'B2', 'B3', 'B4', 'B5', 'C2', 'C3', 'C4', 'C5', 'D2', 'D3', 'D4', 'D5']:
        ws2048[i].font = Font(name='Verdana', size=10)
        ws2048[i].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    # Controls and game status fonts
    for i in ['E2', 'E3', 'E4', 'E5', 'A6']:
        ws2048[i].font = Font(name='Verdana', size=8)

    # Input headings and controls
    ws2048['A1'] = "2048"
    ws2048['C1'] = "Moves:"
    ws2048['E1'] = "Controls"
    ws2048['E2'] = "W = Up, S = Down"
    ws2048['E3'] = "A = Left, D = Right"
    ws2048['E4'] = "R = Reset"
    ws2048['E5'] = "Q = Quit"
    blankCells = ['B1', 'F1', 'F2', 'F3', 'F4', 'F5', 'F6', 'A6', 'B6', 'C6', 'D6', 'E6', 'G6', 'G5', 'G4', 'G3', 'G2', 'G1']
    for i in range(0, len(blankCells)):
        ws2048[blankCells[i]] = ""

    # Save workbook
    wb.save("2048.xlsx")


def createValuesWorkbook():
    # Create workbook for storing values
    wbValues = openpyxl.Workbook()
    wbValues.save("values.xlsx")
